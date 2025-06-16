import pandas as pd
from openpyxl import load_workbook
from openpyxl.styles import PatternFill
from openpyxl.utils.dataframe import dataframe_to_rows

# Load the Excel file and read data
df = pd.read_excel("data.xlsx", sheet_name="Input", header=None)
print("Data loaded from 'Input' sheet:")


df["Average"] = df.mean(axis=1)
print(df)
print("Calculated averages and added to DataFrame.")

wb = load_workbook("data.xlsx")
print("Workbook loaded successfully.")


if "Output" in wb.sheetnames:
    del wb["Output"]
print("Removed existing 'Output' sheet if it was present.")

ws = wb.create_sheet("Output")

# Write DataFrame to sheet
for r_idx, row in enumerate(dataframe_to_rows(df, index=False, header=False), 1):
    for c_idx, value in enumerate(row, 1):
        cell = ws.cell(row=r_idx, column=c_idx, value=value)

        # If this is a value in the "Average" column (last column), apply color
        if c_idx == df.shape[1]:
            # Normalize value between 0–1 based on min and max of averages
            min_val = df["Average"].min()
            max_val = df["Average"].max()
            normalized = (value - min_val) / (max_val - min_val) if max_val != min_val else 0

            # Convert normalized value to RGB from green to red
            red = int(255 * normalized)
            green = int(255 * (1 - normalized))
            color_hex = f"{red:02X}{green:02X}00"

            fill = PatternFill(start_color=color_hex, end_color=color_hex, fill_type="solid")
            cell.fill = fill

# Save updated file
wb.save("data.xlsx")
