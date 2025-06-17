import pandas as pd
from openpyxl import load_workbook
'''
This script reads an existing Excel file, writes messages into a specific sheet of that workbook, and then appends new messages
to an existing template text file. It uses pandas for DataFrame manipulation and openpyxl for Excel file.
Preserve the first 10 lines
Preserve the last 12 lines
Insert generated messages in between
'''
# === STEP 1: Load the existing Excel file into a DataFrame using pandas ===
# Define the path to your existing Excel file
input_file = "value_messages_excel.xlsx"

# Read the 'Input' sheet into a DataFrame
# We use `index_col=0` because the first column contains row labels like "Row1", "Row2", etc.
df = pd.read_excel(input_file, sheet_name="Input", index_col=0)

# Print the data to confirm it's loaded correctly
print("\n Loaded Table:")
print(df)

# === STEP 2: Load the workbook and target sheet using openpyxl ===
# Load the full Excel workbook (preserves formatting and multiple sheets)
wb = load_workbook(input_file)

# Check if the 'Output' sheet already exists
if "Output" not in wb.sheetnames:
    # If not, raise an error so we don't accidentally write somewhere else
    raise ValueError("Sheet 'Output' does not exist in the file.")

# Select the existing 'Output' sheet (don't create or delete anything)
ws_output = wb["Output"]

# === STEP 3: Write messages into the Output sheet ===
# Start writing messages from the first row
row_index = 1

print("\n Writing messages to existing 'Output' sheet...")

# Loop through every row and column of the DataFrame
for row_name in df.index:
    for col_name in df.columns:
        # Get the value at the intersection of current row and column
        value = df.loc[row_name, col_name]

        # Create the message using f-string formatting
        message = f"The value from row: {row_name}, and column: {col_name}, is: {value}"

        # Write the message into column A of the 'Output' sheet
        ws_output.cell(row=row_index, column=1, value=message)

        # Print to terminal (optional debugging)
        print(message)

        # Move to the next row in the sheet
        row_index += 1

print(f"\n Messages written to existing sheet and saved in: '{input_file}'")

# === STEP 4: Save changes ===
# Save the modified workbook, overwriting the original file
wb.save(input_file)

# === STEP 5: Read existing messages from a sample text file and append new messages ===
# Read the existing content from the output file
# This assumes the file is a text file with messages, not an Excel file
with open("output_messages.ids", "r") as file:
    lines = file.readlines()


# Split the preserved parts
top = lines[:10]
bottom = lines[-12:]

print("\n Writing messages to existing text/ids file...")
# Generate new messages
middle = []
for row_name in df.index:
    for col_name in df.columns:
        value = df.loc[row_name, col_name]
        message = f"          The value from row: {row_name}, and column: {col_name}, is: {value}"
        middle.append(message + "\n")

# Combine all parts
updated_lines = top + middle + bottom   ### check the top and bottom parts to see if they should be preserved or they need some changes???

# Write everything back to the file
with open("output_messages.ids", "w") as file:
    file.writelines(updated_lines)
    print(f"\n Success: Messages written to existing output file called: '{file.name}'\n")
